"""
Estimator Module
"""

import os
from openpyxl import load_workbook

from config import TEMPLATE_FILE
from modules.mapper import FieldMapper


class EstimateGenerator:

    def __init__(self):
        self.workbook = load_workbook(TEMPLATE_FILE)
        self.mapper = FieldMapper()

    def write_value(self, sheet_name, cell, value):
        sheet = self.workbook[sheet_name]
        sheet[cell] = value

    def populate_sheet(self, sheet_name, record):

        mapping = self.mapper.get_sheet_mapping(sheet_name)

        for _, row in mapping.iterrows():

            if str(row["Data Source"]).strip() != "ODK":
                continue

            odk_field = str(row["ODK Field"]).strip()

            if odk_field not in record:
                continue

            value = record[odk_field]

            if row["Cell"] == "C7":
                print("Writing C7:", value)

            self.write_value(
                row["Workbook Sheet"],
                row["Cell"],
                value
            )

    def write_fixed_values(self, record):
        # Kharif extent from ODK
        extent = record.get("whs-extent_kharif", "")

        # Format extent
        try:
            extent = round(float(extent), 2)
        except Exception:
            pass

        # Name of Work
        self.write_value(
            "Input Data Sheet-G",
            "C10",
            "Rejuvenation of Water Harvesting Structure"
        )

        # Expected Outcome
        outcome = (
            f"Assured irrigation to {extent} acres of land;\n"
            f"Additional income of ₹25,000 to ₹40,000 per acre;\n"
            f"Ecological development in the village through agro-ecological farming practices."
        )

        self.write_value(
            "Input Data Sheet-G",
            "C11",
            outcome
        )
        self.write_value(
            "Input Data Sheet-T",
            "C66",
            "Kothavalasa"
        )

        self.write_value(
            "Input Data Sheet-T",
            "C70",
            "Mamidipalli"
        )
    def populate_gwr_repeat(self, repeat_records):
        """
        Write GWR repeat records into the Repeat Details sheet.
        """

        sheet = self.workbook["Repeat Details"]

        # Start writing from row 2
        output_row = 2

        # GWR CSV filename from ODK ZIP export
        gwr_filename = "2.Rejuvenation_works-gwr_.csv"

        if gwr_filename not in repeat_records:
            return 2

        gwr_df = repeat_records[gwr_filename]

        for record_no, (_, record) in enumerate(
            gwr_df.iterrows(), start=1
        ):

            # Repeat Group
            sheet.cell(output_row, 2).value = "GWR"

            # Parameter / Work
            sheet.cell(output_row, 3).value = "Guide Wall Repair"

            # Record No
            sheet.cell(output_row, 4).value = record_no

            # Side
            sheet.cell(output_row, 5).value = record.get(
                "gwr_side", ""
            )

            # Chainage From
            sheet.cell(output_row, 6).value = record.get(
                "chainage_gwr_from", ""
            )

            # Chainage To
            sheet.cell(output_row, 7).value = record.get(
                "chainage_gwr_to", ""
            )

            # Length = Chainage To - Chainage From
            sheet.cell(output_row, 8).value = (
                f'=IF(AND(F{output_row}<>"",G{output_row}<>""),'
                f'G{output_row}-F{output_row},"")'
            )

            output_row += 1
        return output_row
    
    def populate_ncg_repeat(self, repeat_records, start_row=2):
        """
        Write NCG repeat records into Repeat Details
        and populate NCG totals/chainages in Input Data Sheet-T.
        """
    
        sheet = self.workbook["Repeat Details"]
        target_sheet = self.workbook["Input Data Sheet-T"]
    
        ncg_filename = "2.Rejuvenation_works-ncg_.csv"
    
        if ncg_filename not in repeat_records:
            return start_row
    
        ncg_df = repeat_records[ncg_filename]
    
        output_row = start_row
    
        left_lengths = []
        right_lengths = []
    
        left_from = []
        left_to = []
    
        right_from = []
        right_to = []
    
        for record_no, (_, record) in enumerate(
            ncg_df.iterrows(), start=1
        ):
    
            side = str(record.get("guidewalls_side", "")).strip().lower()
    
            chainage_from = record.get("chainage_ncg_from", "")
            chainage_to = record.get("chainage_ncg_to", "")
    
            sheet.cell(output_row, 2).value = "NCG"
            sheet.cell(output_row, 3).value = "New Canal Guidewall"
            sheet.cell(output_row, 4).value = record_no
            sheet.cell(output_row, 5).value = side
            sheet.cell(output_row, 6).value = chainage_from
            sheet.cell(output_row, 7).value = chainage_to
    
            # Length = Chainage To - Chainage From
            sheet.cell(output_row, 8).value = (
                f'=IF(AND(F{output_row}<>"",G{output_row}<>""),'
                f'G{output_row}-F{output_row},"")'
            )
    
            # Collect values for Sheet-T
            try:
                cf = float(chainage_from)
                ct = float(chainage_to)
                length = ct - cf
    
                if side == "left":
                    left_lengths.append(length)
                    left_from.append(str(chainage_from))
                    left_to.append(str(chainage_to))
    
                elif side == "right":
                    right_lengths.append(length)
                    right_from.append(str(chainage_from))
                    right_to.append(str(chainage_to))
    
            except (ValueError, TypeError):
                pass
    
            output_row += 1
    
        # -------------------------------------------------
        # Populate Input Data Sheet-T
        # -------------------------------------------------
    
        # LEFT NCG -> Row 19
        if left_lengths:
            target_sheet["E19"] = sum(left_lengths)
            target_sheet["C19"] = "; ".join(left_from)
            target_sheet["D19"] = "; ".join(left_to)
        else:
            target_sheet["E19"] = 0
            target_sheet["C19"] = ""
            target_sheet["D19"] = ""
    
        # RIGHT NCG -> Row 20
        if right_lengths:
            target_sheet["E20"] = sum(right_lengths)
            target_sheet["C20"] = "; ".join(right_from)
            target_sheet["D20"] = "; ".join(right_to)
        else:
            target_sheet["E20"] = 0
            target_sheet["C20"] = ""
            target_sheet["D20"] = ""
    
        return output_row

    def populate_cghi_repeat(self, repeat_records, start_row=2):
        """
        Write Canal Guidewall Height Increase repeat records
        into Repeat Details.
        """
    
        sheet = self.workbook["Repeat Details"]
    
        cghi_filename = "2.Rejuvenation_works-Canal_guidewall_height_increase_.csv"
    
        if cghi_filename not in repeat_records:
            return start_row
    
        cghi_df = repeat_records[cghi_filename]
    
        output_row = start_row
    
        for record_no, (_, record) in enumerate(
            cghi_df.iterrows(), start=1
        ):
    
            sheet.cell(output_row, 2).value = "CGHI"
            sheet.cell(output_row, 3).value = "Increasing height of guide wall"
            sheet.cell(output_row, 4).value = record_no
            sheet.cell(output_row, 5).value = record.get(
                "canal_guidewall_height_increase_side",
                ""
            )

            sheet.cell(output_row, 6).value = record.get(
                "chainage_canal_guidewall_height_increase_from",
                ""
            )
    
            sheet.cell(output_row, 7).value = record.get(
                "chainage_canal_guidewall_height_increase_to",
                ""
            )
    
            # Length = Chainage To - Chainage From
            sheet.cell(output_row, 8).value = (
                f'=IF(AND(F{output_row}<>"",G{output_row}<>""),'
                f'G{output_row}-F{output_row},"")'
            )
    
            output_row += 1
    
        return output_row
    def populate_gwbjl_repeat(self, repeat_records, start_row=2):
        """
        Write GWBJL repeat records into Repeat Details
        and populate E50/E51 in Input Data Sheet-T.
    
        Leak 1 -> E50
        Leak 2 -> E51
        """
    
        repeat_sheet = self.workbook["Repeat Details"]
        target_sheet = self.workbook["Input Data Sheet-T"]
    
        gwbjl_filename = "2.Rejuvenation_works-gwbjl_.csv"
    
        if gwbjl_filename not in repeat_records:
            return start_row
    
        gwbjl_df = repeat_records[gwbjl_filename]
    
        output_row = start_row
    
        leak1_total = 0
        leak2_total = 0
    
        for record_no, (_, record) in enumerate(
            gwbjl_df.iterrows(), start=1
        ):
    
            # ---------------------------------------------
            # Common information
            # ---------------------------------------------
    
            repeat_sheet.cell(output_row, 2).value = "GWBJL"
            repeat_sheet.cell(
                output_row, 3
            ).value = "Hunch through bed and guidewall joint"
    
            repeat_sheet.cell(output_row, 4).value = record_no
    
            # ---------------------------------------------
            # Chainage
            # ---------------------------------------------
    
            chainage_from = record.get(
                "leak1_gwbjl_chinage_from", ""
            )
    
            chainage_to = record.get(
                "leak1_gwbjl_chinage_to", ""
            )
    
            repeat_sheet.cell(output_row, 6).value = chainage_from
            repeat_sheet.cell(output_row, 7).value = chainage_to
    
            # ---------------------------------------------
            # Leak 1
            # ---------------------------------------------
    
            leak1 = record.get(
                "leakage_canal_length_gwbjl_leak1",
                ""
            )
    
            try:
                leak1_value = float(leak1) if leak1 not in ["", None] else 0
            except (ValueError, TypeError):
                leak1_value = 0
    
            leak1_total += leak1_value
    
            repeat_sheet.cell(
                output_row, 8
            ).value = leak1
    
            output_row += 1
    
            # ---------------------------------------------
            # Leak 2
            # ---------------------------------------------
    
            leak2 = record.get(
                "leakage_canal_length_gwbjl_leak2",
                ""
            )
    
            try:
                leak2_value = float(leak2) if leak2 not in ["", None] else 0
            except (ValueError, TypeError):
                leak2_value = 0
    
            leak2_total += leak2_value
    
            repeat_sheet.cell(output_row, 2).value = "GWBJL"
            repeat_sheet.cell(
                output_row, 3
            ).value = "Hunch through bed and guidewall joint"
    
            repeat_sheet.cell(output_row, 4).value = record_no
            repeat_sheet.cell(output_row, 5).value = "left"
    
            repeat_sheet.cell(
                output_row, 6
            ).value = record.get(
                "leak1_gwbjl_chinage_from", ""
            )
    
            repeat_sheet.cell(
                output_row, 7
            ).value = record.get(
                "leak1_gwbjl_chinage_to", ""
            )
    
            repeat_sheet.cell(
                output_row, 8
            ).value = leak2
    
            output_row += 1
    
        # ---------------------------------------------
        # Populate Sheet-T
        # ---------------------------------------------
    
        target_sheet["E50"] = leak1_total
        target_sheet["E51"] = leak2_total
    
        return output_row
    
    def setup_gwr_formulas(self):
        """
        Create Excel 2019+ compatible formulas that consolidate
        GWR records from Repeat Details into Input Data Sheet-T.
        """

        repeat_sheet = self.workbook["Repeat Details"]
        target_sheet = self.workbook["Input Data Sheet-T"]

        # -------------------------------------------------
        # Helper columns
        # J = Right From
        # K = Right To
        # L = Right Length
        # M = Left From
        # N = Left To
        # O = Left Length
        # -------------------------------------------------

        for row in range(2, 501):

            if row == 2:
                # First GWR record

                repeat_sheet.cell(row, 10).value = (
                    f'=IF(AND($B{row}="GWR",LOWER($E{row})="right"),'
                    f'$F{row},"")'
                )

                repeat_sheet.cell(row, 11).value = (
                    f'=IF(AND($B{row}="GWR",LOWER($E{row})="right"),'
                    f'$G{row},"")'
                )

                repeat_sheet.cell(row, 12).value = (
                    f'=IF(AND($B{row}="GWR",LOWER($E{row})="right"),'
                    f'$H{row},0)'
                )

                repeat_sheet.cell(row, 13).value = (
                    f'=IF(AND($B{row}="GWR",LOWER($E{row})="left"),'
                    f'$F{row},"")'
                )

                repeat_sheet.cell(row, 14).value = (
                    f'=IF(AND($B{row}="GWR",LOWER($E{row})="left"),'
                    f'$G{row},"")'
                )

                repeat_sheet.cell(row, 15).value = (
                    f'=IF(AND($B{row}="GWR",LOWER($E{row})="left"),'
                    f'$H{row},0)'
                )

            else:
                # Keep the previous total when the row belongs
                # to the opposite side, and add when it matches.

                # RIGHT - Chainage From
                repeat_sheet.cell(row, 10).value = (
                    f'=IF(AND($B{row}="GWR",'
                    f'LOWER($E{row})="right"),'
                    f'IF(J{row-1}="",TEXT($F{row},"0.0"),'
                    f'J{row-1}&"; "&TEXT($F{row},"0.0")),'
                    f'J{row-1})'
                )

                # RIGHT - Chainage To
                repeat_sheet.cell(row, 11).value = (
                    f'=IF(AND($B{row}="GWR",'
                    f'LOWER($E{row})="right"),'
                    f'IF(K{row-1}="",TEXT($G{row},"0.0"),'
                    f'K{row-1}&"; "&TEXT($G{row},"0.0")),'
                    f'K{row-1})'
                )

                # RIGHT - Length
                repeat_sheet.cell(row, 12).value = (
                    f'=IF(AND($B{row}="GWR",'
                    f'LOWER($E{row})="right"),'
                    f'L{row-1}+$H{row},'
                    f'L{row-1})'
                )

                # LEFT - Chainage From
                repeat_sheet.cell(row, 13).value = (
                    f'=IF(AND($B{row}="GWR",'
                    f'LOWER($E{row})="left"),'
                    f'IF(M{row-1}="",TEXT($F{row},"0.0"),'
                    f'M{row-1}&"; "&TEXT($F{row},"0.0")),'
                    f'M{row-1})'
                )

                # LEFT - Chainage To
                repeat_sheet.cell(row, 14).value = (
                    f'=IF(AND($B{row}="GWR",'
                    f'LOWER($E{row})="left"),'
                    f'IF(N{row-1}="",TEXT($G{row},"0.0"),'
                    f'N{row-1}&"; "&TEXT($G{row},"0.0")),'
                    f'N{row-1})'
                )

                # LEFT - Length
                repeat_sheet.cell(row, 15).value = (
                    f'=IF(AND($B{row}="GWR",'
                    f'LOWER($E{row})="left"),'
                    f'O{row-1}+$H{row},'
                    f'O{row-1})'
                )
        # -------------------------------------------------        
        # Sheet-T
        # GWR Right = row 40
        # GWR Left  = row 41
        # -------------------------------------------------

        target_sheet["C40"] = "='Repeat Details'!J500"
        target_sheet["D40"] = "='Repeat Details'!K500"
        target_sheet["E40"] = "='Repeat Details'!L500"

        target_sheet["C41"] = "='Repeat Details'!M500"
        target_sheet["D41"] = "='Repeat Details'!N500"
        target_sheet["E41"] = "='Repeat Details'!O500"

        # -------------------------------------------------
        # Hide helper columns
        # -------------------------------------------------

        for column in ["J", "K", "L", "M", "N", "O"]:
            repeat_sheet.column_dimensions[column].hidden = True
    
    def setup_cghi_formulas(self):
        """
        Consolidate CGHI records from Repeat Details
        separately for Right and Left sides.
        """
    
        repeat_sheet = self.workbook["Repeat Details"]
        target_sheet = self.workbook["Input Data Sheet-T"]
    
        # -------------------------------------------------
        # Helper columns
        #
        # P = Right From
        # Q = Right To
        # R = Right Length
        # S = Left From
        # T = Left To
        # U = Left Length
        # -------------------------------------------------
    
        for row in range(2, 501):
    
            if row == 2:
    
                # RIGHT
                repeat_sheet.cell(row, 16).value = (
                    f'=IF(AND($B{row}="CGHI",'
                    f'LOWER($E{row})="right"),'
                    f'TEXT($F{row},"0.0"),"")'
                )
    
                repeat_sheet.cell(row, 17).value = (
                    f'=IF(AND($B{row}="CGHI",'
                    f'LOWER($E{row})="right"),'
                    f'TEXT($G{row},"0.0"),"")'
                )
    
                repeat_sheet.cell(row, 18).value = (
                    f'=IF(AND($B{row}="CGHI",'
                    f'LOWER($E{row})="right"),'
                    f'$H{row},0)'
                )
    
                # LEFT
                repeat_sheet.cell(row, 19).value = (
                    f'=IF(AND($B{row}="CGHI",'
                    f'LOWER($E{row})="left"),'
                    f'TEXT($F{row},"0.0"),"")'
                )
    
                repeat_sheet.cell(row, 20).value = (
                    f'=IF(AND($B{row}="CGHI",'
                    f'LOWER($E{row})="left"),'
                    f'TEXT($G{row},"0.0"),"")'
                )
    
                repeat_sheet.cell(row, 21).value = (
                    f'=IF(AND($B{row}="CGHI",'
                    f'LOWER($E{row})="left"),'
                    f'$H{row},0)'
                )
    
            else:
    
                # RIGHT - From
                repeat_sheet.cell(row, 16).value = (
                    f'=IF(AND($B{row}="CGHI",'
                    f'LOWER($E{row})="right"),'
                    f'IF(P{row-1}="",TEXT($F{row},"0.0"),'
                    f'P{row-1}&"; "&TEXT($F{row},"0.0")),'
                    f'P{row-1})'
                )
    
                # RIGHT - To
                repeat_sheet.cell(row, 17).value = (
                    f'=IF(AND($B{row}="CGHI",'
                    f'LOWER($E{row})="right"),'
                    f'IF(Q{row-1}="",TEXT($G{row},"0.0"),'
                    f'Q{row-1}&"; "&TEXT($G{row},"0.0")),'
                    f'Q{row-1})'
                )
    
                # RIGHT - Length
                repeat_sheet.cell(row, 18).value = (
                    f'=IF(AND($B{row}="CGHI",'
                    f'LOWER($E{row})="right"),'
                    f'R{row-1}+$H{row},'
                    f'R{row-1})'
                )
    
                # LEFT - From
                repeat_sheet.cell(row, 19).value = (
                    f'=IF(AND($B{row}="CGHI",'
                    f'LOWER($E{row})="left"),'
                    f'IF(S{row-1}="",TEXT($F{row},"0.0"),'
                    f'S{row-1}&"; "&TEXT($F{row},"0.0")),'
                    f'S{row-1})'
                )
    
                # LEFT - To
                repeat_sheet.cell(row, 20).value = (
                    f'=IF(AND($B{row}="CGHI",'
                    f'LOWER($E{row})="left"),'
                    f'IF(T{row-1}="",TEXT($G{row},"0.0"),'
                    f'T{row-1}&"; "&TEXT($G{row},"0.0")),'
                    f'T{row-1})'
                )
    
                # LEFT - Length
                repeat_sheet.cell(row, 21).value = (
                    f'=IF(AND($B{row}="CGHI",'
                    f'LOWER($E{row})="left"),'
                    f'U{row-1}+$H{row},'
                    f'U{row-1})'
                )
    
        # -------------------------------------------------
        # Sheet-T
        # -------------------------------------------------
    
        # Right side → Row 42
        target_sheet["C42"] = "='Repeat Details'!P500"
        target_sheet["D42"] = "='Repeat Details'!Q500"
        target_sheet["E42"] = "='Repeat Details'!R500"
    
        # Left side → Row 43
        target_sheet["C43"] = "='Repeat Details'!S500"
        target_sheet["D43"] = "='Repeat Details'!T500"
        target_sheet["E43"] = "='Repeat Details'!U500"
    
        # -------------------------------------------------
        # Hide helper columns
        # -------------------------------------------------
    
        for column in ["P", "Q", "R", "S", "T", "U"]:
            repeat_sheet.column_dimensions[column].hidden = True
    def save(self, filename):

        os.makedirs("output", exist_ok=True)

        # Recalculate Excel formulas when the workbook is opened
        self.workbook.calculation.fullCalcOnLoad = True
        self.workbook.calculation.forceFullCalc = True
        self.workbook.calculation.calcMode = "auto"

        self.workbook.save(filename)
